import logging
import os
from datetime import timedelta

import openpyxl as openpyxl
from celery import shared_task
from django.conf import settings
from django.db import transaction
from django.utils import timezone

from main.utils import send_template_mail


logger = logging.getLogger()


@shared_task
def create_recipients(table_id: int):
    from main.models import EmailTable, Recipient

    table = EmailTable.objects.get(id=table_id)
    table_file = table.table_file
    dataframe = openpyxl.load_workbook(table_file.path)
    dataframe1 = dataframe.active

    seller_email_column = dataframe1._cells[(1,1)]
    if seller_email_column.value != "sellerEmail":
        raise Exception

    recipients = []
    recipients_amount = 0
    with transaction.atomic():
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, 1):
                if not col[row].value or col[row].value == "sellerEmail":
                    continue
                recipients.append(
                    Recipient(
                        email=col[row].value,
                        table_id=table_id,
                    )
                )
                recipients_amount += 1
                if len(recipients) == 1000:
                    Recipient.objects.bulk_create(recipients)
                    recipients = []
        Recipient.objects.bulk_create(recipients)
        # EmailTable.objects.filter(table_id=table_id).update(status=EmailTable.ProcessStatus.recipientS_CREATED)
        # initiate_msg_sending()
        EmailTable.objects.filter(id=table_id).update(
            status=EmailTable.ProcessStatus.recipientS_CREATED.value,
            recipients_amount=recipients_amount
        )

    initiate_msg_sending.apply_async((table_id,))


@shared_task
def initiate_msg_sending(table_id: int):
    from main.models import Recipient, EmailTable

    LIMIT = 1000
    OFFSET = 0
    recipients = Recipient.objects.filter(table_id=table_id).distinct("email")[OFFSET: LIMIT]
    EmailTable.objects.filter(id=table_id).update(status=EmailTable.ProcessStatus.MAILING_IN_PROGRESS.value, )
    while recipients:
        for recipient in recipients:
            send_msg.apply_async((recipient.email, table_id))
        OFFSET += LIMIT

        recipients = Recipient.objects.all().distinct("email")[OFFSET: LIMIT + OFFSET]


@shared_task
def send_msg(recipient: str, table_id: int):
    from main.models import EmailTemplate, EmailTable, ExtraFileEmailTable

    template = EmailTemplate.objects.get().template.path
    context = {"docs": []}
    attachments = ExtraFileEmailTable.objects.filter(email_table_id=table_id)
    for attachment in attachments:
        context["docs"].append(attachment.file.file)
    try:
        send_template_mail("новые предложения", template, context, recipient)
    except Exception as e:
        logger.error(f"recipient: {recipient}, table_id: {table_id}, {e}")
        email_table = EmailTable.objects.filter(id=table_id).first()
        if email_table:
            if email_table.sendings + 1 == email_table.recipients_amount:
                email_table.status = EmailTable.ProcessStatus.FINISHED.value
            email_table.sendings += 1
            email_table.save()
        raise e

    email_table = EmailTable.objects.filter(id=table_id).first()
    if email_table:
        if email_table.sendings + 1 == email_table.recipients_amount:
            email_table.status = EmailTable.ProcessStatus.FINISHED.value
        email_table.sendings += 1
        email_table.successful_sendings += 1
        email_table.save()


@shared_task(name="Удаление таблиц (Старше месяца)")
def delete_old_tables():
    from main.models import EmailTable
    tables = EmailTable.objects.filter(uploaded_at__lte=timezone.now() - timedelta(days=30))
    for table in tables:
        if os.path.isfile(table.table_file.path):
            os.remove(table.table_file.path)
        table.delete()


@shared_task(name="Удаление старых файлов")
def delete_old_files():
    from main.models import EmailTemplate, EmailTable, ExtraFile
    template = EmailTemplate.objects.all().values_list("template", flat=True)
    tables = EmailTable.objects.all().values_list("table_file", flat=True)
    extra = ExtraFile.objects.all().values_list("file", flat=True)
    actual_filenames = []
    for file in template:
        actual_filenames.append(file)
    for file in tables:
        actual_filenames.append(file)
    for file in extra:
        actual_filenames.append(file)
    actual_filenames = set(actual_filenames)
    files = os.listdir(settings.MEDIA_ROOT)
    files = set(files)
    files.difference_update(actual_filenames)
    for file in files:
        os.remove(f"{settings.MEDIA_ROOT}/{file}")
